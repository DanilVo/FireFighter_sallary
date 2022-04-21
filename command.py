from time import strftime
import tkinter as tk
from openpyxl import *
import openpyxl

def new_user_data_sheet():
    user_info = openpyxl.load_workbook("User information.xlsx")
    ws = user_info.create_sheet()
    ws.title = f'User {row_number}'
    ws['A1'] = 'Start Shift'
    ws['B1'] = 'End Shift'
    ws['C1'] = 'Total Shift Time'
    user_info.save('User information.xlsx')
    user_info.close()

# refers to to (submite_but) to add data
def registration_data():

    global row_number
    book = openpyxl.load_workbook("Data.xlsx")
    sheet = book.active
    # build in info of datas we recive
    sheet['A1'] = 'User ID'
    sheet['B1'] = 'User Name'
    sheet['C1'] = 'User Password'
    sheet['D1'] = 'First Name'
    sheet['E1'] = 'Last Name'
    sheet['F1'] = 'Registration Date'

    #(sheet.insert_rows() - inserting new row, by default to the top)(row +1 - in serting to place i want) 
    row_number = sheet.max_row
    sheet.insert_rows(row_number + 1)
    sheet[row_number+1][0].value = sheet.max_row  # ID the user will get
    sheet[row_number+1][1].value = name_entry.get()
    sheet[row_number+1][2].value = password_entry.get()
    sheet[row_number+1][3].value = first_name_entry.get()
    sheet[row_number+1][4].value = last_name_entry.get()
    sheet[row_number+1][5].value = strftime("%c")
    
    new_user_data_sheet()

    print('Data writed')
    book.save('Data.xlsx')
    book.close()

    
    # book.create_sheet(f'{name_entry.get()}',0)
    
    

def new_user():

    global name_entry
    global password_entry
    global first_name_entry
    global last_name_entry

    register_window = tk.Tk()
    register_window.title('registration')
    register_window.geometry('300x400')

    name_lable = tk.Label(register_window, text="enter your user:")
    name_lable.grid(row=0,column=0)
    name_entry = tk.Entry(register_window)
    name_entry.grid(row=0,column=1)
    
    password_lable = tk.Label(register_window, text="enter your password:")
    password_lable.grid(row=1,column=0)
    password_entry = tk.Entry(register_window)
    password_entry.grid(row=1,column=1)

    first_name_lable = tk.Label(register_window, text="enter your first name:")
    first_name_lable.grid(row=2,column=0)
    first_name_entry = tk.Entry(register_window)
    first_name_entry.grid(row=2,column=1)

    last_name_lable = tk.Label(register_window, text="enter your last name:")
    last_name_lable.grid(row=3,column=0)
    last_name_entry = tk.Entry(register_window)
    last_name_entry.grid(row=3,column=1)

    submite_but = tk.Button(register_window,text='Save new user!', command= registration_data)
    submite_but.grid(row=4,column=0)

    register_window.mainloop()

